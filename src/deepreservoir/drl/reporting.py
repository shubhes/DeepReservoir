"""deepreservoir.drl.reporting

Build presentation-friendly Excel dashboards from per-evaluation metrics CSV files.

This module is intentionally separate from metrics computation. The raw CSV/JSON
artifacts remain the source of truth; this module scans one or more run folders
for ``eval_metrics.csv`` files and compiles a human-readable workbook intended
for quick comparison, screenshots, and slide copy/paste.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Literal

import json
import math
import re

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover - dependency/import guard
    raise ImportError(
        "deepreservoir.drl.reporting requires openpyxl. Install it in the active environment."
    ) from exc

from deepreservoir.drl.metrics import METRIC_DEFINITIONS, compute_historic_summary_metrics


ColorRule = Literal["none", "higher_good", "lower_good", "zero_best"]


@dataclass(frozen=True)
class MetricDisplaySpec:
    raw_name: str
    label: str
    group: str
    order: int
    include_summary: bool
    include_full: bool
    number_format: str
    color_rule: ColorRule
    description: str = ""


@dataclass(frozen=True)
class EvalMetricsRecord:
    experiment_name: str
    eval_label: str
    display_name: str
    metrics_path: str
    eval_dir: str
    manifest_path: str | None
    last_modified: str
    metadata: dict[str, object]
    metrics: dict[str, float]


# -----------------------------------------------------------------------------
# Public API
# -----------------------------------------------------------------------------


def build_master_metrics_workbook(
    *,
    runs_root: Path | str,
    outpath: Path | str | None = None,
    metrics_filename: str = "eval_metrics.csv",
) -> dict[str, object]:
    """Scan ``runs_root`` for evaluation metrics and build a dashboard workbook.

    Parameters
    ----------
    runs_root
        Root directory containing experiment subdirectories.
    outpath
        Output workbook path. Defaults to ``<runs_root>/master_metrics.xlsx``.
    metrics_filename
        Metrics filename to scan for recursively.

    Returns
    -------
    dict
        Summary information including output path and discovered counts.
    """
    runs_root = Path(runs_root)
    if not runs_root.exists():
        raise FileNotFoundError(f"Runs root does not exist: {runs_root}")
    if outpath is None:
        outpath = runs_root / "master_metrics.xlsx"
    outpath = Path(outpath)

    records = discover_eval_metrics(runs_root, metrics_filename=metrics_filename)
    if not records:
        raise FileNotFoundError(
            f"No {metrics_filename!r} files were found under {runs_root}"
        )

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    summary_df, full_df, index_df, config_eval_df, config_overall_df = _records_to_dataframes(records)
    historic_df = _build_historic_summary_df(records=records, runs_root=runs_root)
    present_metrics = [c for c in full_df.columns if c not in _ID_COLUMNS]
    specs = {name: get_metric_display_spec(name) for name in present_metrics}

    _write_dashboard_sheet(
        wb=wb,
        sheet_name="Summary",
        df=summary_df,
        specs={name: specs[name] for name in summary_df.columns if name not in _ID_COLUMNS},
        title="DeepReservoir core metrics summary",
        subtitle=f"Built from {len(records)} evaluation result(s) under {runs_root}",
    )
    if not historic_df.empty:
        _write_dashboard_sheet(
            wb=wb,
            sheet_name="Historic",
            df=historic_df,
            specs={name: specs[name] for name in historic_df.columns if name != "eval"},
            title="Historic benchmark over eval window(s)",
            subtitle="Computed from historical series over the same evaluation period(s).",
            id_columns=["eval"],
        )
    _write_dashboard_sheet(
        wb=wb,
        sheet_name="Full Dashboard",
        df=full_df,
        specs=specs,
        title="DeepReservoir full metrics dashboard",
        subtitle=f"Built from {len(records)} evaluation result(s) under {runs_root}",
    )
    config_eval_specs = _build_rollup_specs(config_eval_df)
    config_overall_specs = _build_rollup_specs(config_overall_df)

    _write_dashboard_sheet(
        wb=wb,
        sheet_name="Config by Eval",
        df=config_eval_df,
        specs=config_eval_specs,
        title="DeepReservoir config rollup by eval window",
        subtitle="One row per config + eval; metrics are aggregated across repeated seed runs.",
        id_columns=_CONFIG_EVAL_ID_COLUMNS,
    )
    _write_dashboard_sheet(
        wb=wb,
        sheet_name="Config Overall",
        df=config_overall_df,
        specs=config_overall_specs,
        title="DeepReservoir config rollup across all evals",
        subtitle="One row per config; metrics are aggregated across all available eval windows and repeated seed runs.",
        id_columns=_CONFIG_OVERALL_ID_COLUMNS,
    )
    _write_run_index_sheet(wb=wb, df=index_df)
    _write_definitions_sheet(wb=wb, metric_names=present_metrics, specs=specs)

    outpath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(outpath)

    return {
        "outpath": outpath,
        "n_evals": len(records),
        "n_experiments": len({r.experiment_name for r in records}),
    }



def discover_eval_metrics(
    runs_root: Path | str,
    *,
    metrics_filename: str = "eval_metrics.csv",
) -> list[EvalMetricsRecord]:
    """Discover and load evaluation metrics under ``runs_root``."""
    runs_root = Path(runs_root)
    paths = sorted(runs_root.rglob(metrics_filename))
    if not paths:
        return []

    experiment_counts: dict[str, int] = {}
    rel_paths: dict[Path, Path] = {}
    manifest_paths: dict[Path, Path | None] = {}
    experiment_names: dict[Path, str] = {}
    for path in paths:
        rel = path.relative_to(runs_root)
        rel_paths[path] = rel
        manifest_path = _find_manifest_path(runs_root=runs_root, metrics_path=path)
        manifest_paths[path] = manifest_path
        experiment = _experiment_name_from_paths(runs_root=runs_root, rel_metrics_path=rel, manifest_path=manifest_path)
        experiment_names[path] = experiment
        experiment_counts[experiment] = experiment_counts.get(experiment, 0) + 1

    records: list[EvalMetricsRecord] = []
    for path in paths:
        rel = rel_paths[path]
        experiment = experiment_names[path]
        eval_suffix = _relative_eval_label(rel)
        display_name = _build_display_name(
            experiment_name=experiment,
            eval_label=eval_suffix,
            n_evals_for_experiment=experiment_counts.get(experiment, 1),
        )
        manifest_path = manifest_paths[path]
        metadata = _load_manifest_metadata(manifest_path)
        metrics = _load_metrics_row(path)
        records.append(
            EvalMetricsRecord(
                experiment_name=experiment,
                eval_label=eval_suffix,
                display_name=display_name,
                metrics_path=str(rel).replace("\\", "/"),
                eval_dir=str(rel.parent).replace("\\", "/"),
                manifest_path=(None if manifest_path is None else str(manifest_path.relative_to(runs_root)).replace("\\", "/")),
                last_modified=pd.Timestamp(path.stat().st_mtime, unit="s").strftime("%Y-%m-%d %H:%M"),
                metadata=metadata,
                metrics=metrics,
            )
        )

    records.sort(key=lambda r: (r.experiment_name.lower(), r.eval_label.lower(), r.metrics_path.lower()))
    return records


# -----------------------------------------------------------------------------
# Data assembly
# -----------------------------------------------------------------------------


_ID_COLUMNS = ["experiment", "eval", "display_name"]
_CONFIG_EVAL_ID_COLUMNS = ["config_id", "eval", "config_label"]
_CONFIG_OVERALL_ID_COLUMNS = ["config_id", "config_label"]

# Metrics intentionally omitted from the workbook dashboard even if they exist in
# the raw eval_metrics.csv artifacts.
_MASTER_METRICS_EXCLUDE = {
}

# Keep the Excel "Summary" sheet aligned with the compact experiment scoreboard
# exposed by metrics group ``core``. This deliberately excludes reward totals and
# deeper diagnostics so the front page remains a clean cross-run comparison view.
_SUMMARY_METRICS = (
    "dam_safety_frac_days_within_storage_bounds",
    "storage_frac_of_max_possible",
    "esa_min_flow_frac_days_met",
    "flooding_frac_days_met",
    "spr_freq_years_meeting_10000cfs_5d",
    "spr_freq_years_meeting_8000cfs_19d",
    "spr_freq_years_meeting_5000cfs_20d",
    "spr_freq_years_meeting_2500cfs_10d",
    "hydropower_frac_of_max_possible",
    "niip_frac_days_demand_met_in_window",
    "niip_annual_volume_frac_of_contract",
)

_GROUP_ORDER = {
    "Run": 0,
    "Config": 5,
    "Rewards": 10,
    "Dam Safety": 20,
    "ESA / Flow": 30,
    "Flooding": 40,
    "SPR Curve": 50,
    "SPR Thresholds": 60,
    "SPR Targets": 65,
    "Hydropower": 70,
    "NIIP": 80,
    "Reward Components": 90,
    "Other": 100,
}


_SPR_TARGET_FREQUENCIES = {
    "spr_freq_years_meeting_10000cfs_5d": 0.20,
    "spr_freq_years_meeting_8000cfs_19d": 0.33,
    "spr_freq_years_meeting_5000cfs_20d": 0.50,
    "spr_freq_years_meeting_2500cfs_10d": 0.80,
}


def _records_to_dataframes(
    records: list[EvalMetricsRecord],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows: list[dict[str, object]] = []
    index_rows: list[dict[str, object]] = []
    rollup_source_rows: list[dict[str, object]] = []

    for rec in records:
        row: dict[str, object] = {
            "experiment": rec.experiment_name,
            "eval": rec.eval_label,
            "display_name": rec.display_name,
        }
        row.update(rec.metrics)
        rows.append(row)

        idx_row = {
            "experiment": rec.experiment_name,
            "eval": rec.eval_label,
            "display_name": rec.display_name,
            "metrics_path": rec.metrics_path,
            "eval_dir": rec.eval_dir,
            "manifest_path": rec.manifest_path or "",
            "last_modified": rec.last_modified,
            "reward_spec": rec.metadata.get("reward_spec", ""),
            "seed": rec.metadata.get("seed", ""),
            "train_start": rec.metadata.get("train_start", ""),
            "train_end": rec.metadata.get("train_end", ""),
            "val_start": rec.metadata.get("val_start", ""),
            "val_end": rec.metadata.get("val_end", ""),
            "total_timesteps": rec.metadata.get("total_timesteps", ""),
            "gamma": rec.metadata.get("gamma", ""),
            "n_steps": rec.metadata.get("n_steps", ""),
            "batch_size": rec.metadata.get("batch_size", ""),
            "n_epochs": rec.metadata.get("n_epochs", ""),
            "algo": rec.metadata.get("algo", ""),
        }
        index_rows.append(idx_row)

        roll_row: dict[str, object] = {
            "experiment": rec.experiment_name,
            "eval": rec.eval_label,
            "seed": rec.metadata.get("seed", ""),
            "reward_spec": rec.metadata.get("reward_spec", ""),
            "train_start": rec.metadata.get("train_start", ""),
            "train_end": rec.metadata.get("train_end", ""),
            "val_start": rec.metadata.get("val_start", ""),
            "val_end": rec.metadata.get("val_end", ""),
            "total_timesteps": rec.metadata.get("total_timesteps", ""),
            "gamma": rec.metadata.get("gamma", ""),
            "n_steps": rec.metadata.get("n_steps", ""),
            "batch_size": rec.metadata.get("batch_size", ""),
            "n_epochs": rec.metadata.get("n_epochs", ""),
            "algo": rec.metadata.get("algo", ""),
            "episode_length_train": rec.metadata.get("episode_length_train", ""),
            "n_envs": rec.metadata.get("n_envs", ""),
            "val_freq": rec.metadata.get("val_freq", ""),
        }
        for metric_name in _SUMMARY_METRICS:
            roll_row[metric_name] = rec.metrics.get(metric_name, float("nan"))
        rollup_source_rows.append(roll_row)

    df_all = pd.DataFrame(rows)
    metric_cols = [c for c in df_all.columns if c not in _ID_COLUMNS and c not in _MASTER_METRICS_EXCLUDE]
    ordered_metric_cols = _order_metric_columns(metric_cols)
    df_all = df_all[_ID_COLUMNS + ordered_metric_cols]

    summary_cols = _order_metric_columns([c for c in _SUMMARY_METRICS if c in df_all.columns])
    df_summary = df_all[["experiment", "eval"] + summary_cols].copy()
    df_full = df_all.copy()
    df_index = pd.DataFrame(index_rows)

    config_eval_df, config_overall_df = _build_config_rollup_tables(pd.DataFrame(rollup_source_rows))

    return df_summary, df_full, df_index, config_eval_df, config_overall_df


def _build_historic_summary_df(*, records: list[EvalMetricsRecord], runs_root: Path) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    seen: set[str] = set()
    for rec in records:
        eval_label = rec.eval_label or "eval"
        if eval_label in seen:
            continue
        rollout_path = runs_root / rec.eval_dir / "eval_rollout.parquet"
        if not rollout_path.exists():
            continue
        try:
            df_roll = pd.read_parquet(rollout_path)
        except Exception:
            continue
        if "date" in df_roll.columns:
            try:
                df_roll["date"] = pd.to_datetime(df_roll["date"])
                df_roll = df_roll.set_index("date")
            except Exception:
                pass
        if not isinstance(df_roll.index, pd.DatetimeIndex):
            try:
                df_roll.index = pd.to_datetime(df_roll.index)
            except Exception:
                continue
        metrics = compute_historic_summary_metrics(df_roll)
        row: dict[str, object] = {"eval": eval_label}
        ordered_summary_cols = _order_metric_columns(list(_SUMMARY_METRICS))
        for col in ordered_summary_cols:
            row[col] = metrics.get(col, float("nan"))
        rows.append(row)
        seen.add(eval_label)
    ordered_summary_cols = _order_metric_columns(list(_SUMMARY_METRICS))
    return pd.DataFrame(rows, columns=["eval", *ordered_summary_cols]) if rows else pd.DataFrame(columns=["eval", *ordered_summary_cols])


# -----------------------------------------------------------------------------
# Metric display catalog
# -----------------------------------------------------------------------------


_EXACT_SPECS: dict[str, MetricDisplaySpec] = {
    "total_reward": MetricDisplaySpec(
        raw_name="total_reward",
        label="Total reward",
        group="Rewards",
        order=10,
        include_summary=True,
        include_full=True,
        number_format="# ,##0.00".replace(" ", ""),
        color_rule="higher_good",
        description=METRIC_DEFINITIONS.get("total_reward", ""),
    ),
    "mean_reward": MetricDisplaySpec(
        raw_name="mean_reward",
        label="Mean reward",
        group="Rewards",
        order=20,
        include_summary=False,
        include_full=True,
        number_format="0.000",
        color_rule="higher_good",
        description=METRIC_DEFINITIONS.get("mean_reward", ""),
    ),
    "dam_safety_frac_days_within_storage_bounds": MetricDisplaySpec(
        raw_name="dam_safety_frac_days_within_storage_bounds",
        label="Storage in bounds",
        group="Dam Safety",
        order=10,
        include_summary=True,
        include_full=True,
        number_format="0.0%",
        color_rule="higher_good",
        description=METRIC_DEFINITIONS.get("dam_safety_frac_days_within_storage_bounds", ""),
    ),
    "dam_safety_frac_days_below_min_storage": MetricDisplaySpec(
        raw_name="dam_safety_frac_days_below_min_storage",
        label="Below min storage",
        group="Dam Safety",
        order=20,
        include_summary=False,
        include_full=True,
        number_format="0.0%",
        color_rule="lower_good",
        description=METRIC_DEFINITIONS.get("dam_safety_frac_days_below_min_storage", ""),
    ),
    "dam_safety_frac_days_above_max_storage": MetricDisplaySpec(
        raw_name="dam_safety_frac_days_above_max_storage",
        label="Above max storage",
        group="Dam Safety",
        order=30,
        include_summary=False,
        include_full=True,
        number_format="0.0%",
        color_rule="lower_good",
        description=METRIC_DEFINITIONS.get("dam_safety_frac_days_above_max_storage", ""),
    ),
    "dam_safety_max_storage_range_water_year_af": MetricDisplaySpec(
        raw_name="dam_safety_max_storage_range_water_year_af",
        label="Max WY storage range",
        group="Dam Safety",
        order=40,
        include_summary=False,
        include_full=True,
        number_format="# ,##0".replace(" ", ""),
        color_rule="none",
        description=METRIC_DEFINITIONS.get("dam_safety_max_storage_range_water_year_af", ""),
    ),
    "esa_min_flow_frac_days_met": MetricDisplaySpec(
        raw_name="esa_min_flow_frac_days_met",
        label="ESA flow met",
        group="ESA / Flow",
        order=10,
        include_summary=True,
        include_full=True,
        number_format="0.0%",
        color_rule="higher_good",
        description=METRIC_DEFINITIONS.get("esa_min_flow_frac_days_met", ""),
    ),
    "flooding_frac_days_met": MetricDisplaySpec(
        raw_name="flooding_frac_days_met",
        label="Flood-safe days",
        group="Flooding",
        order=10,
        include_summary=True,
        include_full=True,
        number_format="0.0%",
        color_rule="higher_good",
        description=METRIC_DEFINITIONS.get("flooding_frac_days_met", ""),
    ),
    "spr_curve_mean_abs_error_cfs": MetricDisplaySpec(
        raw_name="spr_curve_mean_abs_error_cfs",
        label="SPR curve MAE",
        group="SPR Curve",
        order=10,
        include_summary=False,
        include_full=True,
        number_format="# ,##0.0".replace(" ", ""),
        color_rule="lower_good",
        description=METRIC_DEFINITIONS.get("spr_curve_mean_abs_error_cfs", ""),
    ),
    "spr_curve_mean_error_cfs": MetricDisplaySpec(
        raw_name="spr_curve_mean_error_cfs",
        label="SPR curve mean err",
        group="SPR Curve",
        order=20,
        include_summary=False,
        include_full=True,
        number_format="# ,##0.0;[Red](# ,##0.0)".replace(" ", ""),
        color_rule="zero_best",
        description=METRIC_DEFINITIONS.get("spr_curve_mean_error_cfs", ""),
    ),
    "spr_curve_frac_days_within_500cfs": MetricDisplaySpec(
        raw_name="spr_curve_frac_days_within_500cfs",
        label="SPR within 500",
        group="SPR Curve",
        order=30,
        include_summary=True,
        include_full=True,
        number_format="0.0%",
        color_rule="higher_good",
        description=METRIC_DEFINITIONS.get("spr_curve_frac_days_within_500cfs", ""),
    ),
    "hydropower_frac_of_max_possible": MetricDisplaySpec(
        raw_name="hydropower_frac_of_max_possible",
        label="Hydro / max possible",
        group="Hydropower",
        order=10,
        include_summary=True,
        include_full=True,
        number_format="0.0%",
        color_rule="higher_good",
        description=METRIC_DEFINITIONS.get("hydropower_frac_of_max_possible", ""),
    ),
    "storage_frac_of_max_possible": MetricDisplaySpec(
        raw_name="storage_frac_of_max_possible",
        label="Storage / max possible",
        group="Dam Safety",
        order=15,
        include_summary=True,
        include_full=True,
        number_format="0.0%",
        color_rule="higher_good",
        description=METRIC_DEFINITIONS.get("storage_frac_of_max_possible", ""),
    ),
    "niip_frac_days_demand_met_in_window": MetricDisplaySpec(
        raw_name="niip_frac_days_demand_met_in_window",
        label="NIIP demand met",
        group="NIIP",
        order=10,
        include_summary=True,
        include_full=True,
        number_format="0.0%",
        color_rule="higher_good",
        description=METRIC_DEFINITIONS.get("niip_frac_days_demand_met_in_window", ""),
    ),
    "niip_annual_volume_frac_of_contract": MetricDisplaySpec(
        raw_name="niip_annual_volume_frac_of_contract",
        label="NIIP vol / contract",
        group="NIIP",
        order=20,
        include_summary=True,
        include_full=True,
        number_format="0.0%",
        color_rule="higher_good",
        description=METRIC_DEFINITIONS.get("niip_annual_volume_frac_of_contract", ""),
    ),
}


def get_metric_display_spec(raw_name: str) -> MetricDisplaySpec:
    if raw_name in _EXACT_SPECS:
        return _EXACT_SPECS[raw_name]

    m = re.fullmatch(r"spr_freq_years_meeting_(\d+)cfs_(\d+)d", raw_name)
    if m:
        thr, dur = int(m.group(1)), int(m.group(2))
        return MetricDisplaySpec(
            raw_name=raw_name,
            label=f"{_thr_label(thr)} / {dur}d freq",
            group="SPR Targets",
            order=100 + _spr_order(thr, dur),
            include_summary=False,
            include_full=True,
            number_format="0.0%",
            color_rule="none",
            description=_definition_for_pattern(raw_name),
        )

    m = re.fullmatch(r"spr_target_frequency_(\d+)cfs_(\d+)d", raw_name)
    if m:
        thr, dur = int(m.group(1)), int(m.group(2))
        return MetricDisplaySpec(
            raw_name=raw_name,
            label=f"{_thr_label(thr)} / {dur}d target",
            group="SPR Targets",
            order=110 + _spr_order(thr, dur),
            include_summary=False,
            include_full=True,
            number_format="0.0%",
            color_rule="none",
            description=_definition_for_pattern(raw_name),
        )

    m = re.fullmatch(r"spr_overachievement_(\d+)cfs_(\d+)d", raw_name)
    if m:
        thr, dur = int(m.group(1)), int(m.group(2))
        return MetricDisplaySpec(
            raw_name=raw_name,
            label=f"{_thr_label(thr)} / {dur}d Δfreq",
            group="SPR Targets",
            order=120 + _spr_order(thr, dur),
            include_summary=(raw_name in _SUMMARY_METRICS),
            include_full=True,
            number_format="0.0%;[Red](0.0%)",
            color_rule="higher_good",
            description=_definition_for_pattern(raw_name),
        )

    m = re.fullmatch(r"spr_mean_max_consec_days_(\d+)cfs", raw_name)
    if m:
        thr = int(m.group(1))
        return MetricDisplaySpec(
            raw_name=raw_name,
            label=f"{_thr_label(thr)} total days above",
            group="SPR Thresholds",
            order=200 + _spr_order(thr, 0),
            include_summary=False,
            include_full=True,
            number_format="0.0",
            color_rule="higher_good",
            description=_definition_for_pattern(raw_name),
        )

    m = re.fullmatch(r"spr_mean_frac_window_days_above_(\d+)cfs", raw_name)
    if m:
        thr = int(m.group(1))
        return MetricDisplaySpec(
            raw_name=raw_name,
            label=f"{_thr_label(thr)} frac above",
            group="SPR Thresholds",
            order=210 + _spr_order(thr, 0),
            include_summary=False,
            include_full=True,
            number_format="0.0%",
            color_rule="higher_good",
            description=_definition_for_pattern(raw_name),
        )

    m = re.fullmatch(r"sum_rc_(.+)", raw_name)
    if m:
        name = _humanize_metric_token(m.group(1))
        return MetricDisplaySpec(
            raw_name=raw_name,
            label=f"Σ {name}",
            group="Reward Components",
            order=1000 + _stable_order_key(raw_name),
            include_summary=False,
            include_full=True,
            number_format="0.000",
            color_rule="none",
            description=_definition_for_pattern(raw_name),
        )

    m = re.fullmatch(r"mean_rc_(.+)", raw_name)
    if m:
        name = _humanize_metric_token(m.group(1))
        return MetricDisplaySpec(
            raw_name=raw_name,
            label=f"Mean {name}",
            group="Reward Components",
            order=1100 + _stable_order_key(raw_name),
            include_summary=False,
            include_full=True,
            number_format="0.000",
            color_rule="none",
            description=_definition_for_pattern(raw_name),
        )

    return MetricDisplaySpec(
        raw_name=raw_name,
        label=_humanize_metric_name(raw_name),
        group="Other",
        order=2000 + _stable_order_key(raw_name),
        include_summary=False,
        include_full=True,
        number_format=_guess_number_format(raw_name),
        color_rule=_guess_color_rule(raw_name),
        description=_definition_for_pattern(raw_name),
    )


_ROLLUP_METADATA_SPECS: dict[str, MetricDisplaySpec] = {
    "reward_spec": MetricDisplaySpec("reward_spec", "Reward spec", "Config", 10, False, True, "@", "none"),
    "train_start": MetricDisplaySpec("train_start", "Train start", "Config", 20, False, True, "@", "none"),
    "train_end": MetricDisplaySpec("train_end", "Train end", "Config", 30, False, True, "@", "none"),
    "val_start": MetricDisplaySpec("val_start", "Val start", "Config", 40, False, True, "@", "none"),
    "val_end": MetricDisplaySpec("val_end", "Val end", "Config", 50, False, True, "@", "none"),
    "total_timesteps": MetricDisplaySpec("total_timesteps", "Timesteps", "Config", 60, False, True, "#,##0", "none"),
    "gamma": MetricDisplaySpec("gamma", "Gamma", "Config", 70, False, True, "0.000", "none"),
    "n_steps": MetricDisplaySpec("n_steps", "n_steps", "Config", 80, False, True, "#,##0", "none"),
    "batch_size": MetricDisplaySpec("batch_size", "Batch size", "Config", 90, False, True, "#,##0", "none"),
    "n_epochs": MetricDisplaySpec("n_epochs", "n_epochs", "Config", 100, False, True, "#,##0", "none"),
    "n_runs": MetricDisplaySpec("n_runs", "Runs", "Config", 110, False, True, "0", "none"),
    "n_seeds": MetricDisplaySpec("n_seeds", "Seeds", "Config", 120, False, True, "0", "none"),
    "n_experiments": MetricDisplaySpec("n_experiments", "Experiments", "Config", 130, False, True, "0", "none"),
    "n_evals": MetricDisplaySpec("n_evals", "Eval windows", "Config", 140, False, True, "0", "none"),
}


def _build_rollup_specs(df: pd.DataFrame) -> dict[str, MetricDisplaySpec]:
    specs: dict[str, MetricDisplaySpec] = {}
    id_columns = set(_CONFIG_EVAL_ID_COLUMNS) | set(_CONFIG_OVERALL_ID_COLUMNS)
    for name in df.columns:
        if name in id_columns:
            continue
        if name in _ROLLUP_METADATA_SPECS:
            specs[name] = _ROLLUP_METADATA_SPECS[name]
            continue
        mean_m = re.fullmatch(r"mean__(.+)", name)
        if mean_m:
            base = get_metric_display_spec(mean_m.group(1))
            specs[name] = MetricDisplaySpec(
                raw_name=name,
                label=f"μ {base.label}",
                group=base.group,
                order=(base.order * 10) + 1,
                include_summary=False,
                include_full=True,
                number_format=base.number_format,
                color_rule=base.color_rule,
                description=f"Mean across grouped runs for {base.label}.",
            )
            continue
        std_m = re.fullmatch(r"std__(.+)", name)
        if std_m:
            base = get_metric_display_spec(std_m.group(1))
            specs[name] = MetricDisplaySpec(
                raw_name=name,
                label=f"σ {base.label}",
                group=base.group,
                order=(base.order * 10) + 2,
                include_summary=False,
                include_full=True,
                number_format=base.number_format,
                color_rule="lower_good",
                description=f"Across-run standard deviation for {base.label}.",
            )
            continue
        specs[name] = MetricDisplaySpec(
            raw_name=name,
            label=_humanize_metric_name(name),
            group="Config",
            order=900,
            include_summary=False,
            include_full=True,
            number_format=_guess_number_format(name),
            color_rule="none",
            description="",
        )
    return specs


def _first_non_empty(values: Iterable[object]) -> object:
    for val in values:
        if pd.isna(val):
            continue
        sval = str(val).strip()
        if sval and sval.lower() not in {"nan", "none"}:
            return val
    return ""


def _count_unique_non_empty(values: Iterable[object]) -> int:
    cleaned = {
        str(v).strip()
        for v in values
        if not pd.isna(v) and str(v).strip() and str(v).strip().lower() not in {"nan", "none"}
    }
    return int(len(cleaned))


def _config_label_from_experiments(values: Iterable[object]) -> str:
    cleaned = [
        str(v).strip()
        for v in values
        if not pd.isna(v) and str(v).strip() and str(v).strip().lower() not in {"nan", "none"}
    ]
    if not cleaned:
        return ""
    stripped = [re.sub(r"(?:[_\-]+)?seed[_\-]*\d+$", "", name, flags=re.IGNORECASE).rstrip("_- ") for name in cleaned]
    unique = sorted({name or orig for name, orig in zip(stripped, cleaned)})
    if len(unique) == 1:
        return unique[0]
    return unique[0]


def _std0(values: pd.Series) -> float:
    vals = pd.to_numeric(values, errors="coerce").dropna()
    if vals.empty:
        return float("nan")
    return float(vals.std(ddof=0))


def _make_config_signature(row: pd.Series) -> str:
    payload = {
        "algo": row.get("algo", ""),
        "reward_spec": row.get("reward_spec", ""),
        "train_start": row.get("train_start", ""),
        "train_end": row.get("train_end", ""),
        "val_start": row.get("val_start", ""),
        "val_end": row.get("val_end", ""),
        "total_timesteps": row.get("total_timesteps", ""),
        "gamma": row.get("gamma", ""),
        "n_steps": row.get("n_steps", ""),
        "batch_size": row.get("batch_size", ""),
        "n_epochs": row.get("n_epochs", ""),
        "episode_length_train": row.get("episode_length_train", ""),
        "n_envs": row.get("n_envs", ""),
        "val_freq": row.get("val_freq", ""),
    }
    return json.dumps(payload, sort_keys=True, default=str)


_ROLLUP_META_COLS = [
    "reward_spec",
    "train_start",
    "train_end",
    "val_start",
    "val_end",
    "total_timesteps",
    "gamma",
    "n_steps",
    "batch_size",
    "n_epochs",
]


def _build_config_rollup_tables(source_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    metric_cols = sorted(
        [c for c in _SUMMARY_METRICS if c in source_df.columns],
        key=lambda name: (_GROUP_ORDER.get(get_metric_display_spec(name).group, 999), get_metric_display_spec(name).order, get_metric_display_spec(name).label.lower(), name.lower()),
    )
    if source_df.empty:
        return pd.DataFrame(columns=_CONFIG_EVAL_ID_COLUMNS), pd.DataFrame(columns=_CONFIG_OVERALL_ID_COLUMNS)

    work = source_df.copy()
    work["config_signature"] = work.apply(_make_config_signature, axis=1)

    config_meta = (
        work.groupby("config_signature", dropna=False)
        .agg(
            config_label=("experiment", _config_label_from_experiments),
            reward_spec=("reward_spec", _first_non_empty),
            train_start=("train_start", _first_non_empty),
            train_end=("train_end", _first_non_empty),
            val_start=("val_start", _first_non_empty),
            val_end=("val_end", _first_non_empty),
            total_timesteps=("total_timesteps", _first_non_empty),
            gamma=("gamma", _first_non_empty),
            n_steps=("n_steps", _first_non_empty),
            batch_size=("batch_size", _first_non_empty),
            n_epochs=("n_epochs", _first_non_empty),
        )
        .reset_index()
        .sort_values(by=["reward_spec", "train_start", "train_end", "val_start", "val_end", "total_timesteps", "gamma", "n_steps", "batch_size", "n_epochs", "config_label"], kind="stable")
        .reset_index(drop=True)
    )
    config_meta["config_id"] = [f"cfg_{i:03d}" for i in range(1, len(config_meta) + 1)]

    work = work.merge(config_meta[["config_signature", "config_id", "config_label"]], on="config_signature", how="left")

    named_aggs_overall: dict[str, tuple[str, object]] = {
        "config_label": ("config_label", _first_non_empty),
        "n_runs": ("experiment", "size"),
        "n_seeds": ("seed", _count_unique_non_empty),
        "n_experiments": ("experiment", _count_unique_non_empty),
        "n_evals": ("eval", _count_unique_non_empty),
    }
    named_aggs_by_eval: dict[str, tuple[str, object]] = {
        "config_label": ("config_label", _first_non_empty),
        "n_runs": ("experiment", "size"),
        "n_seeds": ("seed", _count_unique_non_empty),
        "n_experiments": ("experiment", _count_unique_non_empty),
    }
    for col in _ROLLUP_META_COLS:
        named_aggs_overall[col] = (col, _first_non_empty)
        named_aggs_by_eval[col] = (col, _first_non_empty)
    for metric in metric_cols:
        named_aggs_overall[f"mean__{metric}"] = (metric, "mean")
        named_aggs_overall[f"std__{metric}"] = (metric, _std0)
        named_aggs_by_eval[f"mean__{metric}"] = (metric, "mean")
        named_aggs_by_eval[f"std__{metric}"] = (metric, _std0)

    overall = (
        work.groupby(["config_id", "config_signature"], dropna=False)
        .agg(**named_aggs_overall)
        .reset_index()
        .drop(columns=["config_signature"])
    )
    by_eval = (
        work.groupby(["config_id", "config_signature", "eval"], dropna=False)
        .agg(**named_aggs_by_eval)
        .reset_index()
        .drop(columns=["config_signature"])
    )

    overall_cols = _CONFIG_OVERALL_ID_COLUMNS + _ROLLUP_META_COLS + ["n_runs", "n_seeds", "n_experiments", "n_evals"]
    by_eval_cols = _CONFIG_EVAL_ID_COLUMNS + _ROLLUP_META_COLS + ["n_runs", "n_seeds", "n_experiments"]
    metric_rollup_cols = [item for metric in metric_cols for item in (f"mean__{metric}", f"std__{metric}")]
    overall = overall[overall_cols + metric_rollup_cols]
    by_eval = by_eval[by_eval_cols + metric_rollup_cols]
    return by_eval, overall


# -----------------------------------------------------------------------------
# Workbook writing
# -----------------------------------------------------------------------------


_THIN_GRAY = Side(style="thin", color="D9D9D9")
_BORDER = Border(bottom=_THIN_GRAY)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F1F1F")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SUBHEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_SUBTITLE_FONT = Font(color="666666", italic=True)
_MISSING_FILL = PatternFill(fill_type="solid", fgColor="E6E6E6")
_NEUTRAL_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
_ID_FILL = PatternFill(fill_type="solid", fgColor="F7F7F7")
_GROUP_FILLS = {
    "Run": "2F5597",
    "Config": "3D85C6",
    "Rewards": "7F6000",
    "Dam Safety": "5B9BD5",
    "ESA / Flow": "70AD47",
    "Flooding": "C55A11",
    "SPR Curve": "8064A2",
    "SPR Thresholds": "7030A0",
    "SPR Targets": "8E7CC3",
    "Hydropower": "1F4E78",
    "NIIP": "BF9000",
    "Reward Components": "595959",
    "Other": "7F7F7F",
}


def _write_dashboard_sheet(
    *,
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    specs: dict[str, MetricDisplaySpec],
    title: str,
    subtitle: str,
    id_columns: Iterable[str] | None = None,
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    id_columns = list(id_columns or _ID_COLUMNS)
    id_column_set = set(id_columns)

    # Title block
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = _SUBTITLE_FONT

    ordered_cols = list(df.columns)
    groups = _dashboard_group_sequence(ordered_cols, specs, id_columns=id_columns)

    header_row_1 = 4
    header_row_2 = 5
    data_start_row = 6

    # Group header row
    for group, start_idx, end_idx in groups:
        col1 = get_column_letter(start_idx)
        col2 = get_column_letter(end_idx)
        ws.merge_cells(f"{col1}{header_row_1}:{col2}{header_row_1}")
        c = ws.cell(row=header_row_1, column=start_idx, value=group)
        c.fill = PatternFill(fill_type="solid", fgColor=_GROUP_FILLS.get(group, "7F7F7F"))
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Column header row
    id_column_set = set(id_columns or _ID_COLUMNS)
    for col_idx, name in enumerate(ordered_cols, start=1):
        label = name if name in id_column_set else specs[name].label
        cell = ws.cell(row=header_row_2, column=col_idx, value=label)
        group = "Run" if name in id_column_set else specs[name].group
        cell.fill = PatternFill(fill_type="solid", fgColor=_GROUP_FILLS.get(group, "595959"))
        cell.font = _SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=data_start_row):
        max_lines = 1
        for col_idx, name in enumerate(ordered_cols, start=1):
            val = row[name]
            if name == "reward_spec" and not pd.isna(val):
                sval = str(val)
                parts = [p.strip() for p in sval.split(",") if p.strip()]
                val = "\n".join(parts) if parts else sval
                max_lines = max(max_lines, max(1, str(val).count("\n") + 1))
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.isna(val):
                cell.value = None
            else:
                cell.value = val.item() if hasattr(val, "item") else val
            wrap = name in {"reward_spec", "metrics_path", "manifest_path"}
            cell.alignment = Alignment(horizontal=("left" if name in id_column_set or wrap else "center"), vertical="center", wrap_text=wrap)
            cell.border = _BORDER
            if name in id_column_set:
                cell.fill = _ID_FILL
            else:
                cell.number_format = specs[name].number_format
        if max_lines > 1:
            ws.row_dimensions[row_idx].height = max(18, 15 * max_lines)

    # Color metric cells after values are written.
    for name in ordered_cols:
        if name in id_column_set:
            continue
        col_idx = ordered_cols.index(name) + 1
        col_values = pd.to_numeric(df[name], errors="coerce")
        spr_target = _spr_target_for_metric_name(name)
        if spr_target is not None:
            fills = _fills_for_threshold(col_values, threshold=spr_target)
        else:
            fills = _fills_for_series(col_values, specs[name].color_rule, metric_name=name)
        for i, fill in enumerate(fills, start=data_start_row):
            ws.cell(row=i, column=col_idx).fill = fill

    freeze_col = get_column_letter(min(len(ordered_cols) + 1, len(id_columns) + 1)) if ordered_cols else "A"
    ws.freeze_panes = f"{freeze_col}{data_start_row}"
    ws.auto_filter.ref = f"A{header_row_2}:{get_column_letter(len(ordered_cols))}{max(data_start_row, ws.max_row)}"
    ws.row_dimensions[header_row_1].height = 22
    ws.row_dimensions[header_row_2].height = 34
    _set_column_widths(ws, ordered_cols, specs, id_columns=id_columns)



def _write_run_index_sheet(*, wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Run Index")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Run index"
    ws["A1"].font = _TITLE_FONT

    header_row = 3
    for col_idx, name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for row_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        max_lines = 1
        for col_idx, name in enumerate(df.columns, start=1):
            val = row[name]
            if name == "reward_spec" and not pd.isna(val):
                sval = str(val)
                parts = [p.strip() for p in sval.split(",") if p.strip()]
                val = "\n".join(parts) if parts else sval
                max_lines = max(max_lines, max(1, str(val).count("\n") + 1))
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.isna(val):
                cell.value = None
            else:
                cell.value = val.item() if hasattr(val, "item") else val
            wrap = name in {"reward_spec", "metrics_path", "manifest_path"}
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
            cell.border = _BORDER
            if col_idx <= 3:
                cell.fill = _ID_FILL
        if max_lines > 1:
            ws.row_dimensions[row_idx].height = max(18, 15 * max_lines)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(df.shape[1])}{max(header_row + 1, ws.max_row)}"
    _set_plain_column_widths(ws)
    if "reward_spec" in df.columns:
        reward_col_idx = list(df.columns).index("reward_spec") + 1
        ws.column_dimensions[get_column_letter(reward_col_idx)].width = 72



def _write_definitions_sheet(
    *,
    wb: Workbook,
    metric_names: Iterable[str],
    specs: dict[str, MetricDisplaySpec],
) -> None:
    ws = wb.create_sheet("Definitions")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Metric definitions"
    ws["A1"].font = _TITLE_FONT

    headers = ["group", "display_label", "raw_name", "number_format", "color_rule", "description"]
    header_row = 3
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    ordered = _order_metric_columns(list(metric_names))
    for row_idx, raw_name in enumerate(ordered, start=header_row + 1):
        spec = specs[raw_name]
        values = [
            spec.group,
            spec.label,
            raw_name,
            spec.number_format,
            spec.color_rule,
            spec.description,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = _BORDER

    ws.freeze_panes = "A4"
    _set_plain_column_widths(ws)

# -----------------------------------------------------------------------------
# Helpers: discovery / metadata
# -----------------------------------------------------------------------------


def _load_metrics_row(path: Path) -> dict[str, float]:
    df = pd.read_csv(path)
    if df.empty:
        return {}
    row = df.iloc[0].to_dict()
    out: dict[str, float] = {}
    for key, value in row.items():
        try:
            out[key] = float(value)
        except Exception:
            out[key] = float("nan")
    return out





def _experiment_name_from_paths(*, runs_root: Path, rel_metrics_path: Path, manifest_path: Path | None) -> str:
    if manifest_path is not None:
        try:
            rel_run = manifest_path.parent.relative_to(runs_root)
            if len(rel_run.parts) >= 1:
                return "/".join(rel_run.parts)
        except Exception:
            pass
    return rel_metrics_path.parts[0] if len(rel_metrics_path.parts) >= 1 else rel_metrics_path.parent.name

def _find_manifest_path(*, runs_root: Path, metrics_path: Path) -> Path | None:
    rel = metrics_path.relative_to(runs_root)
    if len(rel.parts) == 0:
        return None
    candidate = runs_root / rel.parts[0] / "run_manifest.json"
    if candidate.exists():
        return candidate
    for parent in [metrics_path.parent] + list(metrics_path.parents):
        if parent == runs_root.parent:
            break
        cand = parent / "run_manifest.json"
        if cand.exists():
            return cand
        if parent == runs_root:
            break
    return None



def _load_manifest_metadata(manifest_path: Path | None) -> dict[str, object]:
    if manifest_path is None or not manifest_path.exists():
        return {}
    try:
        with open(manifest_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return {}

    cfg = data.get("config", {}) if isinstance(data, dict) else {}
    train = cfg.get("train", {}) if isinstance(cfg.get("train"), dict) else {}
    val = cfg.get("val", {}) if isinstance(cfg.get("val"), dict) else {}

    train_invocations = data.get("train_invocations", []) if isinstance(data, dict) else []
    last_inv = train_invocations[-1] if train_invocations else {}
    ppo_args = last_inv.get("ppo_args", {}) if isinstance(last_inv, dict) else {}

    return {
        "reward_spec": cfg.get("reward_spec", ""),
        "seed": cfg.get("seed", ""),
        "algo": cfg.get("algo", ""),
        "gamma": cfg.get("gamma", ""),
        "n_envs": cfg.get("n_envs", ""),
        "episode_length_train": cfg.get("episode_length_train", ""),
        "train_start": train.get("train_start", ""),
        "train_end": train.get("train_end", ""),
        "val_start": val.get("val_start", ""),
        "val_end": val.get("val_end", ""),
        "total_timesteps": last_inv.get("requested_total_timesteps", ""),
        "val_freq": last_inv.get("val_freq", ""),
        "n_steps": ppo_args.get("n_steps", ""),
        "batch_size": ppo_args.get("batch_size", ""),
        "n_epochs": ppo_args.get("n_epochs", ""),
    }



def _relative_eval_label(rel_metrics_path: Path) -> str:
    parts = list(rel_metrics_path.parts[:-1])
    if not parts:
        return "eval"
    parts = parts[1:]
    parts = [p for p in parts if not re.fullmatch(r"seed[_-]*\d+", p, flags=re.IGNORECASE)]
    if not parts:
        return "eval"
    cleaned = [re.sub(r"^eval__", "", p) for p in parts]
    label = "/".join(cleaned).strip("/")
    return label or "eval"



def _build_display_name(*, experiment_name: str, eval_label: str, n_evals_for_experiment: int) -> str:
    normalized = eval_label.strip().strip("/")
    if n_evals_for_experiment == 1 and normalized in {"", "eval", "."}:
        return experiment_name
    if normalized in {"", "."}:
        normalized = "root"
    return f"{experiment_name} | {normalized}"


# -----------------------------------------------------------------------------
# Helpers: ordering / labeling
# -----------------------------------------------------------------------------


def _order_metric_columns(metric_names: list[str]) -> list[str]:
    specs = [get_metric_display_spec(name) for name in metric_names]
    ordered = sorted(
        specs,
        key=lambda s: (_GROUP_ORDER.get(s.group, 999), s.order, s.label.lower(), s.raw_name.lower()),
    )
    return [s.raw_name for s in ordered]



def _dashboard_group_sequence(
    ordered_cols: list[str],
    specs: dict[str, MetricDisplaySpec],
    *,
    id_columns: Iterable[str] | None = None,
) -> list[tuple[str, int, int]]:
    groups: list[tuple[str, int, int]] = []
    id_column_set = set(id_columns or _ID_COLUMNS)
    current_group = None
    start_idx = 1
    for idx, name in enumerate(ordered_cols, start=1):
        group = "Run" if name in id_column_set else specs[name].group
        if current_group is None:
            current_group = group
            start_idx = idx
            continue
        if group != current_group:
            groups.append((current_group, start_idx, idx - 1))
            current_group = group
            start_idx = idx
    if current_group is not None:
        groups.append((current_group, start_idx, len(ordered_cols)))
    return groups



def _thr_label(thr: int) -> str:
    if thr % 1000 == 0:
        return f"{thr // 1000}k"
    return f"{thr:,}".replace(",", "")



def _spr_order(thr: int, dur: int) -> int:
    base = {
        (10000, 5): 10,
        (8000, 19): 20,
        (5000, 20): 30,
        (2500, 10): 40,
        (10000, 0): 50,
        (8000, 0): 60,
        (5000, 0): 70,
        (2500, 0): 80,
    }
    return base.get((thr, dur), 900 + int(abs(thr)) + int(abs(dur)))



def _definition_for_pattern(raw_name: str) -> str:
    if raw_name.startswith("spr_freq_years_meeting_"):
        return METRIC_DEFINITIONS.get("spr_freq_years_meeting_*", "")
    if raw_name.startswith("spr_target_frequency_"):
        return METRIC_DEFINITIONS.get("spr_target_frequency_*", "")
    if raw_name.startswith("spr_overachievement_"):
        return METRIC_DEFINITIONS.get("spr_overachievement_*", "")
    if raw_name.startswith("spr_mean_max_consec_days_"):
        return METRIC_DEFINITIONS.get("spr_mean_max_consec_days_*", "")
    if raw_name.startswith("spr_mean_frac_window_days_above_"):
        return METRIC_DEFINITIONS.get("spr_mean_frac_window_days_above_*", "")
    if raw_name.startswith("sum_rc_"):
        return METRIC_DEFINITIONS.get("sum_rc_*", "")
    if raw_name.startswith("mean_rc_"):
        return METRIC_DEFINITIONS.get("mean_rc_*", "")
    return METRIC_DEFINITIONS.get(raw_name, "")



def _humanize_metric_name(name: str) -> str:
    text = name.replace("_", " ").replace("/", " / ")
    text = re.sub(r"\s+", " ", text).strip()
    return _apply_acronyms(text).title().replace("Niip", "NIIP").replace("Esa", "ESA").replace("Spr", "SPR")



def _humanize_metric_token(name: str) -> str:
    text = name.replace(".", " ").replace("_", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return _apply_acronyms(text)



def _apply_acronyms(text: str) -> str:
    replacements = {
        r"\bniip\b": "NIIP",
        r"\besa\b": "ESA",
        r"\bspr\b": "SPR",
        r"\bcfs\b": "CFS",
        r"\baf\b": "AF",
        r"\bwy\b": "WY",
        r"\bmwh\b": "MWh",
        r"\brc\b": "rc",
    }
    out = text
    for pattern, repl in replacements.items():
        out = re.sub(pattern, repl, out, flags=re.IGNORECASE)
    return out



def _guess_number_format(raw_name: str) -> str:
    if any(tok in raw_name for tok in ("frac_", "_frac", "frequency")):
        return "0.0%"
    if any(tok in raw_name for tok in ("_af", "_cfs", "_mwh")):
        if raw_name.endswith("_af"):
            return "#,##0"
        return "#,##0.0"
    if "days" in raw_name:
        return "0.0"
    if "reward" in raw_name:
        return "0.000"
    return "0.000"



def _guess_color_rule(raw_name: str) -> ColorRule:
    if any(tok in raw_name for tok in ("mean_abs_error", "below_", "above_", "penalty", "capped_or_limited")):
        return "lower_good"
    if "mean_error" in raw_name:
        return "zero_best"
    if any(tok in raw_name for tok in ("frac_", "_frac", "frequency", "reward")):
        return "higher_good"
    return "none"



def _stable_order_key(text: str) -> int:
    return sum((idx + 1) * ord(ch) for idx, ch in enumerate(text))


# -----------------------------------------------------------------------------
# Helpers: formatting / coloring
# -----------------------------------------------------------------------------


def _set_column_widths(
    ws,
    ordered_cols: list[str],
    specs: dict[str, MetricDisplaySpec],
    *,
    id_columns: Iterable[str] | None = None,
) -> None:
    id_column_set = set(id_columns or _ID_COLUMNS)
    for col_idx, name in enumerate(ordered_cols, start=1):
        if name == "experiment":
            width = 24
        elif name == "eval":
            width = 20
        elif name == "display_name":
            width = 32
        elif name == "config_id":
            width = 10
        elif name == "config_label":
            width = 24
        elif name == "reward_spec":
            width = 44
        elif name in {"train_start", "train_end", "val_start", "val_end"}:
            width = 12
        elif name in {"total_timesteps", "n_steps", "batch_size", "n_epochs", "n_runs", "n_seeds", "n_experiments", "n_evals"}:
            width = 12
        elif name == "gamma":
            width = 10
        elif name in id_column_set:
            width = 22
        else:
            width = max(12, min(18, len(specs[name].label) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width



def _set_plain_column_widths(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        width = 18
        if col_idx in (1, 2, 3):
            width = 22
        if col_idx in (4, 5, 6):
            width = 34
        ws.column_dimensions[get_column_letter(col_idx)].width = width



_FIXED_RANGE_RULES: dict[str, tuple[ColorRule, float, float]] = {
    "dam_safety_frac_days_within_storage_bounds": ("higher_good", 0.0, 1.0),
    "esa_min_flow_frac_days_met": ("higher_good", 0.0, 1.0),
    "flooding_frac_days_met": ("higher_good", 0.0, 1.0),
    "spr_curve_mean_abs_error_cfs": ("lower_good", 0.0, 5000.0),
    "spr_curve_mean_error_cfs": ("zero_best", -5000.0, 5000.0),
    "spr_curve_frac_days_within_500cfs": ("higher_good", 0.0, 1.0),
    "spr_freq_years_meeting_10000cfs_5d": ("higher_good", 0.0, 1.0),
    "spr_freq_years_meeting_8000cfs_19d": ("higher_good", 0.0, 1.0),
    "spr_freq_years_meeting_5000cfs_20d": ("higher_good", 0.0, 1.0),
    "spr_freq_years_meeting_2500cfs_10d": ("higher_good", 0.0, 1.0),
    "hydropower_frac_of_max_possible": ("higher_good", 0.0, 1.0),
    "storage_frac_of_max_possible": ("higher_good", 0.0, 1.0),
    "niip_frac_days_demand_met_in_window": ("higher_good", 0.0, 1.0),
    "niip_annual_volume_frac_of_contract": ("higher_good", 0.0, 1.5),
}


def _base_metric_name(name: str) -> str:
    m = re.fullmatch(r"(?:mean|std)__(.+)", name)
    return m.group(1) if m else name


def _score_from_fixed_range(value: float, *, low: float, high: float, rule: ColorRule) -> float:
    if rule in {"higher_good", "lower_good"}:
        span = high - low
        score = 0.5 if span == 0.0 else (float(value) - float(low)) / span
        score = max(0.0, min(1.0, score))
        if rule == "lower_good":
            score = 1.0 - score
        return score
    max_abs = max(abs(float(low)), abs(float(high)))
    if max_abs == 0.0:
        return 1.0
    return max(0.0, 1.0 - min(abs(float(value)) / max_abs, 1.0))


def _spr_target_for_metric_name(name: str) -> float | None:
    return _SPR_TARGET_FREQUENCIES.get(_base_metric_name(name))


def _fills_for_threshold(series: pd.Series, *, threshold: float) -> list[PatternFill]:
    vals = pd.to_numeric(series, errors="coerce")
    fills: list[PatternFill] = []
    for v in vals:
        if pd.isna(v) or not math.isfinite(float(v)):
            fills.append(_MISSING_FILL)
            continue
        fills.append(_fill_for_score(1.0 if float(v) >= float(threshold) else 0.0))
    return fills


def _fills_for_series(series: pd.Series, rule: ColorRule, *, metric_name: str | None = None) -> list[PatternFill]:
    vals = pd.to_numeric(series, errors="coerce")
    finite = [float(v) for v in vals if pd.notna(v) and math.isfinite(float(v))]
    if rule == "none":
        return [_NEUTRAL_FILL if pd.notna(v) else _MISSING_FILL for v in vals]
    if not finite:
        return [_MISSING_FILL for _ in vals]

    fixed = _FIXED_RANGE_RULES.get(_base_metric_name(metric_name or ""))
    fills: list[PatternFill] = []
    if fixed is not None and fixed[0] == rule:
        _, low, high = fixed
        for v in vals:
            if pd.isna(v) or not math.isfinite(float(v)):
                fills.append(_MISSING_FILL)
                continue
            fills.append(_fill_for_score(_score_from_fixed_range(float(v), low=low, high=high, rule=rule)))
        return fills

    if rule in {"higher_good", "lower_good"}:
        vmin = min(finite)
        vmax = max(finite)
        span = vmax - vmin
        for v in vals:
            if pd.isna(v) or not math.isfinite(float(v)):
                fills.append(_MISSING_FILL)
                continue
            score = 0.5 if span == 0.0 else (float(v) - vmin) / span
            if rule == "lower_good":
                score = 1.0 - score
            fills.append(_fill_for_score(score))
        return fills

    max_abs = max(abs(v) for v in finite)
    for v in vals:
        if pd.isna(v) or not math.isfinite(float(v)):
            fills.append(_MISSING_FILL)
            continue
        score = 1.0 if max_abs == 0.0 else 1.0 - min(abs(float(v)) / max_abs, 1.0)
        fills.append(_fill_for_score(score))
    return fills



def _fill_for_score(score: float) -> PatternFill:
    score = max(0.0, min(1.0, float(score)))
    red = (244, 199, 195)
    yellow = (252, 232, 178)
    green = (183, 225, 205)
    if score <= 0.5:
        t = score / 0.5 if score > 0.0 else 0.0
        rgb = _blend_rgb(red, yellow, t)
    else:
        t = (score - 0.5) / 0.5
        rgb = _blend_rgb(yellow, green, t)
    return PatternFill(fill_type="solid", fgColor=_rgb_to_hex(rgb))



def _blend_rgb(a: tuple[int, int, int], b: tuple[int, int, int], t: float) -> tuple[int, int, int]:
    return tuple(int(round(a[i] + (b[i] - a[i]) * t)) for i in range(3))



def _rgb_to_hex(rgb: tuple[int, int, int]) -> str:
    return "{:02X}{:02X}{:02X}".format(*rgb)


